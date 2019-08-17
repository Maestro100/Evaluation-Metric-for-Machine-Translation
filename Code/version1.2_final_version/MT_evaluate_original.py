from Tkinter import *
import Tkinter
import tkMessageBox
from meteor import *
from gtm import *
from bleu_new import *
from wer import *
from amber import *
import openpyxl
from openpyxl import *

def window(main):
    main.title('MT Evaluator')
    main.update_idletasks()
    width=main.winfo_width()
    height=main.winfo_height()
    x=(main.winfo_screenwidth() // 2) - (width // 2)
    y=(main.winfo_screenheight() // 2) - (height // 2)
    main.geometry('{}x{}+{}+{}'.format(width,height,x,y))

def calc(wb):
    sheet=wb.active
    n=sheet.max_row-1

    sheet.cell(row=1,column=3).value='BLEU Score'
    sheet.cell(row=1,column=4).value='meteor Score'
    sheet.cell(row=1,column=5).value='gtm Score'
    sheet.cell(row=1,column=6).value='Amber Score'
    sheet.cell(row=1,column=7).value='Wer Score'

    for i in range(n):
        translated=sheet.cell(row=i+2,column=2).value
        reference=sheet.cell(row=i+2,column=1).value


        bleu_score=Bleu(reference,translated)
        sheet.cell(row=i+2,column=3).value=bleu_score

        meteor_score=meteor(reference,translated)
        sheet.cell(row=i+2,column=4).value=meteor_score

        gtm_score=gtm(reference,translated)
        sheet.cell(row=i+2,column=5).value=gtm_score

        amber_score=amber(reference,translated)
        sheet.cell(row=i+2,column=6).value=amber_score

        wer_score=wer(reference,translated)
        sheet.cell(row=i+2,column=7).value=wer_score

    wb.save('Scores.xlsx')
    tkMessageBox.showinfo("Done !",'New file with scores has been created  \n with the name "Scores.xlsx in the" \n  same directory as the application' )

def new_winF(): # new window definition
    newwin = Toplevel(root)
    newwin.title('Import Excel Sheet')
    #newwin.geometry('{}x{}'.format(400, 200))
    xx=(newwin.winfo_screenwidth() // 2) - 203
    yy=(newwin.winfo_screenheight() // 2) - 87
    newwin.geometry('{}x{}+{}+{}'.format(406,175,xx,yy))
    only_frame = Frame(newwin, bg='#01a5a5', width=406, height=175)
    only_frame.grid(row=0,sticky='ew')
    fname_label1 = Label(only_frame, text="Have an excel file named",font="Georgia 17", bg='#01a5a5')
    fname_label2 = Label(only_frame, text=" 'Sentences.xlsx' ",font="Georgia 24", bg='#01a5a5')
    fname_label3 = Label(only_frame, text="in the same directory as the application",font="Georgia 17", bg='#01a5a5')
    #fname_label4 = Label(only_frame, text="as the application",font="Georgia 17", bg='#e57b7b')
    wb=load_workbook('Sentences.xlsx')
    BB=Button(only_frame, text ="Calculate Scores",font="Times 25", background='#ffffff',fg='#010101' ,command = lambda:calc(wb) )

    #fname_entry = Entry(newwin , background="pink" ,font='Georgia 25', justify='center')
    fname_label1.grid(row=0,column=0, sticky="ew")
    fname_label2.grid(row=1,column=0, sticky="ew")
    fname_label3.grid(row=2,column=0, sticky="ew")
    #fname_label4.grid(row=3,column=0, sticky="ew")
    BB.grid(row=4,column=0, sticky="ew")
    newwin.resizable(False,False)


root = Tk()
root.title('Machine Translation Evaluator')
root.geometry('{}x{}'.format(1250, 680))
root.resizable(False,False)

window(root)

tkMessageBox.showinfo("Instruction",'Enter the Translated and Reference  \n sentences in the respective text boxes \n and click on "Submit" to get the scores' )

# create all of the main containers
top_frame = Frame(root, bg='#ff9999', width=1250, height=120)
left_frame = Frame(root, bg='#ff9999', width=600, height=450)
center_frame = Frame(root,bg='#444444',width=42,height=450)
right_frame = Frame(root, bg='#ff9999', width=600, height=450)
btm_frame = Frame(root, bg='#ff9999', width=1250, height=110)

# la    yout all of the main containers
root.grid_rowconfigure(1, weight=1)
root.grid_columnconfigure(0, weight=1)
top_frame.grid(row=0, sticky="ew")
left_frame.grid(row=1, sticky="nsw")
center_frame.grid(row=1)
right_frame.grid(row=1, sticky="nse")
btm_frame.grid(row=2, sticky="ew")



# create the widgets for the top frame
Title_label = Label(top_frame, text='MT EVALUATOR',font="Georgia 45", bg='#ff9999')
# layout the widgets in the top frame
Title_label.grid(row=0, columnspan=3)
Title_label.place(x=370,y=5)



# create the widgets for the left frame
traslated_label = Label(left_frame, text='Translated sentence',font="Georgia 17", bg='#ff9999')
translated_entry = Text(left_frame,height=8,width=70,background='pink')
reference_label = Label(left_frame, text='Reference sentence',font="Georgia 17", bg='#ff9999')
reference_entry = Text(left_frame,height=8,width=70,background='pink')
# layout the widgets in the left frame
traslated_label.grid(row=0,columnspan=1,pady=25)
translated_entry.grid(row=1,padx=20,pady=5)
reference_label.grid(row=2,pady=15)
reference_entry.grid(row=3,padx=20,pady=5)





# create the widgets for the right frame
meteor_label = Label(right_frame, text='METEOR Score',font ='Georgia 15', bg='#ff9999')
meteor_entry = Entry(right_frame ,background="pink", font='Georgia 25',justify='center' )
wer_label = Label(right_frame, text='WER Score',font ='Georgia 15', bg='#ff9999')
wer_entry = Entry(right_frame , background="pink" ,font='Georgia 25',justify='center')
gtm_label = Label(right_frame, text='GTM Score',font ='Georgia 15', bg='#ff9999')
gtm_entry = Entry(right_frame, background="pink",font='Georgia 25',justify='center')
amber_label = Label(right_frame, text='AMBER Score',font ='Georgia 15', bg='#ff9999')
amber_entry = Entry(right_frame, background="pink",font='Georgia 25',justify='center')
bleu_label = Label(right_frame, text='BLEU Score',font ='Georgia 15', bg='#ff9999')
bleu_entry = Entry(right_frame, background="pink",font='Georgia 25',justify='center')
# layout the widgets in the right frame
meteor_label.grid(row=0,column=0,sticky='w',pady='10px',padx=25)
meteor_entry.grid(row=0,column=1,pady='10px',padx=15)
wer_label.grid(row=1,column=0,sticky='w',pady='20px',padx=25)
wer_entry.grid(row=1,column=1,pady='20px',padx=15)
gtm_label.grid(row=2,column=0,sticky='w',pady='20px',padx=25)
gtm_entry.grid(row=2,column=1,pady='20px',padx=15)
amber_label.grid(row=3,column=0,sticky='w',pady='20px',padx=25)
amber_entry.grid(row=3,column=1,pady='20px',padx=15)
bleu_label.grid(row=4,column=0,sticky='w',pady='20px',padx=25)
bleu_entry.grid(row=4,column=1,pady='20px',padx=15)


menubar = Menu(root)
def hello():
    print "hello!"

def give_about():
    tkMessageBox.showinfo("Version 1.0",'Developed by  \n Arjya Das \n Kishan Kumar \n Ayush Chaurasia ' )

# create a pulldown menu, and add it to the menu bar
filemenu = Menu(menubar, tearoff=0)
#filemenu.add_command(label="Open", command=hello)
#filemenu.add_command(label="Save", command=hello)
#filemenu.add_separator()

filemenu.add_command(label="Import Excel Sheet", command=new_winF)
#filemenu.add_command(label="Open new windows", command=hello)

filemenu.add_separator()
filemenu.add_command(label="Exit", command=root.quit)
menubar.add_cascade(label="File", menu=filemenu)

helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label="About", command=give_about)
menubar.add_cascade(label="Help", menu=helpmenu)
# create more pulldown menus
'''
editmenu = Menu(menubar, tearoff=0)
editmenu.add_command(label="Cut", command=hello)
editmenu.add_command(label="Copy", command=hello)
editmenu.add_command(label="Paste", command=hello)
menubar.add_cascade(label="Edit", menu=editmenu)


helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label="About", command=hello)
menubar.add_cascade(label="Help", menu=helpmenu)
'''

# display the menu
root.config(menu=menubar)


#Fnunctions
def process():
    hypot = translated_entry.get("1.0",'end-1c').lower()
    refer = reference_entry.get("1.0",'end-1c').lower()
    meteor_entry.delete(0,END)
    wer_entry.delete(0,END)
    gtm_entry.delete(0,END)
    amber_entry.delete(0,END)
    bleu_entry.delete(0,END)
    Entry.insert(meteor_entry,0,"%.3f" % meteor(refer, hypot)) #wer takes parameters r and h as strings
    Entry.insert(wer_entry,0,"%.3f" % wer(refer,hypot)) #gtm takes parameters r and h as strings
    Entry.insert(gtm_entry,0,"%.3f" % gtm(refer,hypot)) #bleu takes parameters r and h as strings
    Entry.insert(amber_entry,0,"%.3f" % amber(refer,hypot))
    Entry.insert(bleu_entry,0,"%.3f" % Bleu(refer,hypot))
print()

def reset():
    meteor_entry.delete(0,END)
    wer_entry.delete(0,END)
    gtm_entry.delete(0,END)
    amber_entry.delete(0,END)
    bleu_entry.delete(0,END)


# create the widgets for the bottom frame
B=Button(btm_frame, text ="Submit",font="Times 35", background='#ffa2a2',fg='#010101',command = process )
# layout the widgets in the bottom frame
B.grid(row=0,sticky='e')
B.place(relx=.42, rely=.1)




root.mainloop()
