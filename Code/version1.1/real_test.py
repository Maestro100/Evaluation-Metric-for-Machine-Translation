
import openpyxl
import time
start_time=time.time()
from meteor import *
from gtm import *
from bleu_new import *
from wer import *
from amber import *
from openpyxl import *
wb=load_workbook("all_the_scores_bleu.xlsx")
sheet=wb.active
n=sheet.max_row-1

for i in range(n):
    translated=sheet.cell(row=i+2,column=3).value
    reference=sheet.cell(row=i+2,column=2).value

    meteor_score=meteor(reference,translated)
    sheet.cell(row=i+2,column=5).value=meteor_score

    #bleu_score=Bleu(reference,translated)
    #sheet.cell(row=i+2,column=4).value=bleu_score

    #gtm_score=gtm(reference,translated)
    #sheet.cell(row=i+2,column=6).value=gtm_score
    #wer_score=wer(reference,translated)
    #sheet.cell(row=i+2,column=8).value=wer_score

    #amber_score=amber(reference,translated)
    #sheet.cell(row=i+2,column=7).value=amber_score

print "total time required = "
print time.time()-start_time
wb.save("bleu_and_meteor.xlsx")
